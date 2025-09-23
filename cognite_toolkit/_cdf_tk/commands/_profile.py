import importlib.util
import itertools
from abc import ABC, abstractmethod
from collections import defaultdict
from collections.abc import Callable, Hashable, Iterable, Mapping
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from functools import cached_property, partial
from itertools import zip_longest
from pathlib import Path
from typing import TYPE_CHECKING, ClassVar, Generic, Literal, TypeAlias, TypeVar, overload
from zipfile import BadZipFile

import questionary
from cognite.client.data_classes import Transformation
from cognite.client.exceptions import CogniteException
from rich import box
from rich.console import Console
from rich.live import Live
from rich.spinner import Spinner
from rich.table import Table

from cognite_toolkit._cdf_tk.client import ToolkitClient
from cognite_toolkit._cdf_tk.client.data_classes.raw import RawProfileResults, RawTable
from cognite_toolkit._cdf_tk.constants import MAX_ROW_ITERATION_RUN_QUERY
from cognite_toolkit._cdf_tk.exceptions import ToolkitMissingDependencyError, ToolkitThrottledError, ToolkitValueError
from cognite_toolkit._cdf_tk.utils.aggregators import (
    AssetAggregator,
    AssetCentricAggregator,
    EventAggregator,
    FileAggregator,
    LabelAggregator,
    LabelCountAggregator,
    MetadataAggregator,
    RelationshipAggregator,
    SequenceAggregator,
    TimeSeriesAggregator,
)
from cognite_toolkit._cdf_tk.utils.cdf import get_transformation_sources, raw_row_count
from cognite_toolkit._cdf_tk.utils.interactive_select import AssetCentricDestinationSelect, AssetInteractiveSelect
from cognite_toolkit._cdf_tk.utils.sql_parser import SQLParser, SQLTable
from cognite_toolkit._cdf_tk.utils.text import sanitize_spreadsheet_title
from cognite_toolkit._cdf_tk.utils.useful_types import AssetCentricDestinationType

from ._base import ToolkitCommand

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


class WaitingAPICallClass:
    def __bool__(self) -> bool:
        return False


WaitingAPICall = WaitingAPICallClass()


PendingCellValue: TypeAlias = int | float | str | bool | None | WaitingAPICallClass
CellValue: TypeAlias = int | float | str | bool | None

T_Index = TypeVar("T_Index", bound=Hashable)


class ProfileCommand(ToolkitCommand, ABC, Generic[T_Index]):
    spreadsheet_max_column_width = 70.0

    def __init__(
        self,
        output_spreadsheet: Path | None = None,
        print_warning: bool = True,
        skip_tracking: bool = False,
        silent: bool = False,
    ) -> None:
        super().__init__(print_warning, skip_tracking, silent)
        self.table_title = self.__class__.__name__.removesuffix("Command")
        self.output_spreadsheet: Path | None = output_spreadsheet
        if output_spreadsheet is not None:
            self._validate_openpyxl_installed()

    class Columns:  # Placeholder for columns, subclasses should define their own Columns class
        ...

    spinner_args: ClassVar[Mapping] = dict(name="arc", text="loading...", style="bold green", speed=1.0)

    max_workers = 8
    is_dynamic_table = False

    @cached_property
    def columns(self) -> tuple[str, ...]:
        return (
            tuple([value for attr, value in self.Columns.__dict__.items() if not attr.startswith("_")])
            if hasattr(self, "Columns")
            else tuple()
        )

    @staticmethod
    def _validate_openpyxl_installed() -> None:
        """Ensure that openpyxl is installed if output_spreadsheet is set."""
        if importlib.util.find_spec("openpyxl") is None:
            raise ToolkitMissingDependencyError(
                "Writing to a spreadsheet requires 'openpyxl'. Install with 'pip install \"cognite-toolkit[table]\"'"
            )

    def create_profile_table(self, client: ToolkitClient, sheet: str | None = None) -> list[dict[str, CellValue]]:
        console = Console()
        with console.status("Setting up", spinner="aesthetic", speed=0.4) as _:
            table = self.create_initial_table(client)
        with (
            Live(self.draw_table(table), refresh_per_second=4, console=console) as live,
            ThreadPoolExecutor(max_workers=self.max_workers) as executor,
        ):
            while True:
                current_calls = {
                    #
                    executor.submit(self.create_api_callable(row, col, client)): (row, col)
                    for (row, col), cell in table.items()
                    if cell is WaitingAPICall
                }
                if not current_calls:
                    break
                for future in as_completed(current_calls):
                    row, col = current_calls[future]
                    try:
                        result = future.result()
                    except CogniteException as e:
                        result = type(e).__name__
                    except ToolkitThrottledError as e:
                        result = f"Throttled: Wait {e.wait_time_seconds:.0f} seconds"
                    table[(row, col)] = self.format_result(result, row, col)
                    if self.is_dynamic_table:
                        table = self.update_table(table, result, row, col)
                    live.update(self.draw_table(table))
        result = self.as_record_format(table, allow_waiting_api_call=False)
        if self.output_spreadsheet is not None:
            self._write_to_spreadsheet(result, list(self.columns), self.output_spreadsheet, sheet=sheet)
        return result

    @abstractmethod
    def create_initial_table(self, client: ToolkitClient) -> dict[tuple[T_Index, str], PendingCellValue]:
        """
        Create the initial table with placeholders for API calls.
        Each cell that requires an API call should be initialized with WaitingAPICall.
        """
        raise NotImplementedError("Subclasses must implement create_initial_table.")

    @abstractmethod
    def create_api_callable(self, row: T_Index, col: str, client: ToolkitClient) -> Callable:
        raise NotImplementedError("Subclasses must implement call_api.")

    def format_result(self, result: object, row: T_Index, col: str) -> CellValue:
        """
        Format the result of an API call for display in the table.
        This can be overridden by subclasses to customize formatting.
        """
        if isinstance(result, int | float | bool | str):
            return result
        raise NotImplementedError("Subclasses must implement format_result.")

    def update_table(
        self,
        current_table: dict[tuple[T_Index, str], PendingCellValue],
        result: object,
        row: T_Index,
        col: str,
    ) -> dict[tuple[T_Index, str], PendingCellValue]:
        raise NotImplementedError("Subclasses must implement update_table.")

    def draw_table(self, table: dict[tuple[T_Index, str], PendingCellValue]) -> Table:
        rich_table = Table(
            title=self.table_title,
            title_justify="left",
            show_header=True,
            header_style="bold magenta",
            box=box.MINIMAL,
        )
        for col in self.columns:
            rich_table.add_column(col)

        rows = self.as_record_format(table)

        last_row: list[str | Spinner] = []
        for row in rows:
            this_row = [self._as_cell(value) for value in row.values()]
            draw_row = self._create_draw_row(this_row, last_row)
            rich_table.add_row(*draw_row)
            last_row = this_row
        return rich_table

    @classmethod
    def _create_draw_row(cls, this_row: list[str | Spinner], last_row: list[str | Spinner]) -> list[str | Spinner]:
        """Creates the row to be drawn. This skips sequential cells that have not changed
        such that the table does not have too many repeated values and thus becomes easier to read.
        """
        draw_row: list[str | Spinner] = []
        row_has_changed = False
        for cell, last_cell in zip_longest(this_row, last_row, fillvalue=""):
            if not row_has_changed and cls._should_skip_drawing(cell, last_cell):
                cell = ""
            else:
                # If the first cell in the row has changed, all remaining cells in the row should be drawn.
                row_has_changed = True
            draw_row.append(cell)
        return draw_row

    @classmethod
    def _should_skip_drawing(cls, cell: str | Spinner, last_cell: str | Spinner) -> bool:
        return cell == last_cell or (isinstance(cell, Spinner) and isinstance(last_cell, Spinner))

    @classmethod
    @overload
    def as_record_format(
        cls, table: dict[tuple[T_Index, str], PendingCellValue], allow_waiting_api_call: Literal[True] = True
    ) -> list[dict[str, PendingCellValue]]: ...

    @classmethod
    @overload
    def as_record_format(
        cls,
        table: dict[tuple[T_Index, str], PendingCellValue],
        allow_waiting_api_call: Literal[False],
    ) -> list[dict[str, CellValue]]: ...

    @classmethod
    def as_record_format(
        cls,
        table: dict[tuple[T_Index, str], PendingCellValue],
        allow_waiting_api_call: bool = True,
    ) -> list[dict[str, PendingCellValue]] | list[dict[str, CellValue]]:
        rows: list[dict[str, PendingCellValue]] = []
        row_indices: dict[T_Index, int] = {}
        for (row, col), value in table.items():
            if value is WaitingAPICall and not allow_waiting_api_call:
                value = None
            if row not in row_indices:
                row_indices[row] = len(rows)
                rows.append({col: value})
            else:
                rows[row_indices[row]][col] = value
        return rows

    def _as_cell(self, value: PendingCellValue) -> str | Spinner:
        if isinstance(value, WaitingAPICallClass):
            return Spinner(**self.spinner_args)
        elif isinstance(value, int):
            return f"{value:,}"
        elif isinstance(value, float):
            return f"{value:.2f}"
        elif value is None:
            return "-"
        return str(value)

    def _write_to_spreadsheet(
        self, data: list[dict[str, CellValue]], columns: list[str], output_spreadsheet: Path, sheet: str | None = None
    ) -> None:
        """Write the profile data to a spreadsheet."""
        # Local import as this is an optional dependency
        from openpyxl import Workbook, load_workbook

        sheet_name = sanitize_spreadsheet_title(sheet or self.table_title)[
            :31
        ]  # Limit title to 31 characters for Excel compatibility
        if output_spreadsheet.exists():
            try:
                workbook = load_workbook(output_spreadsheet)
            except (OSError, BadZipFile) as e:
                raise ToolkitValueError(
                    f"Failed to open {output_spreadsheet.as_posix()!r}. "
                    "Please ensure the file is not corrupted or open in another application."
                ) from e
            if sheet_name in workbook.sheetnames:
                raise ToolkitValueError(f"Sheet '{sheet_name}' already exists in {output_spreadsheet.as_posix()}.")
            worksheet = workbook.create_sheet(title=sheet_name)
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name

        worksheet.append(columns)

        for row in data:
            worksheet.append(list(row.values()))

        self._style_sheet(worksheet, columns)

        try:
            workbook.save(output_spreadsheet)
        except OSError as e:
            raise ToolkitValueError(
                f"Failed to write to {output_spreadsheet.as_posix()!r}. "
                "Please ensure the file is not open, and that you have sufficient permissions and disk space."
            ) from e
        self.console(f"Profile data written to sheet {sheet!r} in {output_spreadsheet.as_posix()!r}")

    def _style_sheet(self, sheet: "Worksheet", columns: list[str]) -> None:
        """Styles the sheet with the given headers.

        Args:
            sheet: The sheet to style.
            headers: The headers to style.
        """
        # Local import as this is an optional dependency
        from openpyxl.cell import MergedCell
        from openpyxl.styles import Font, PatternFill

        # This freezes all rows above the given row
        sheet.freeze_panes = "A2"

        # Make the header row bold, larger, and colored
        for cell, *_ in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(columns)):
            cell.font = Font(bold=True, size=15)
            cell.fill = PatternFill(fgColor="A9DFBF", patternType="solid")
        # Adjust columns width based on widest cell in each column
        for column_cells in sheet.columns:
            try:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
            except ValueError:
                max_length = 0

            # Find first non-merged cell to safely get column letter
            safe_cell = next((c for c in column_cells if not isinstance(c, MergedCell)), None)
            if safe_cell is None:
                continue  # Skip if no such cell is found

            column_letter = safe_cell.column_letter
            current = sheet.column_dimensions[column_letter].width or (max_length + 0.5)
            sheet.column_dimensions[column_letter].width = min(
                max(current, max_length + 0.5), self.spreadsheet_max_column_width
            )

    def _ask_store_file(self) -> None:
        if file_path := questionary.path("Where do you want to save the profile?").ask():
            self.output_spreadsheet = Path(file_path)


@dataclass(frozen=True)
class AssetIndex:
    aggregator: str
    data_set_external_id: str | None = None
    source: RawTable | None = None


class ProfileAssetCommand(ProfileCommand[AssetIndex]):
    def __init__(
        self,
        output_spreadsheet: Path | None = None,
        print_warning: bool = True,
        skip_tracking: bool = False,
        silent: bool = False,
    ) -> None:
        super().__init__(output_spreadsheet, print_warning, skip_tracking, silent)
        self.table_title = "Asset Profile for Hierarchy"
        self.hierarchy: str | None = None
        self.aggregators: dict[str, MetadataAggregator] = {}
        self.profile_row_limit = self.max_profile_row_limit

    class Columns:
        Resource = "Resource"
        Count = "Count"
        DataSets = "DataSet"
        DataSetCount = "DataSet Count"
        Transformations = "Transformation"
        RawTable = "Raw Table"
        RowCount = "Rows"
        ColumnCount = "Columns"

    is_dynamic_table = True
    max_profile_row_limit = 10_000  # The number of rows to profile to get the number of columns.
    # The actual limit is 256 MB of data.
    # Ref https://github.com/cognitedata/profiler-api/blob/main/src/main/scala/com/cognite/raw_profiler/Profile.scala#L37
    profile_timeout_seconds = 60 * 4  # Timeout for the profiling operation in seconds,

    def assets(
        self,
        client: ToolkitClient,
        hierarchy: str | None = None,
        profile_row_limit: int = max_profile_row_limit,
        verbose: bool = False,
    ) -> list[dict[str, CellValue]]:
        """
        Profile assets in the given hierarchy.
        This method will create a table with the count of assets, events, files, timeseries, sequences,
        relationships, and labels in the specified hierarchy.
        """
        if hierarchy is None:
            self.hierarchy = AssetInteractiveSelect(client, "profile").select_hierarchy(allow_empty=False)
            self._ask_store_file()
        else:
            self.hierarchy = hierarchy
        if profile_row_limit <= 0 or profile_row_limit > self.max_profile_row_limit:
            raise ToolkitValueError(
                f"Profile row limit must be between 1 and {self.max_profile_row_limit}, got {profile_row_limit}."
            )
        self.table_title = f"Asset Profile for Hierarchy: {self.hierarchy}"
        self.profile_row_limit = profile_row_limit
        self.aggregators = {
            agg.display_name: agg
            for agg in [
                AssetAggregator(client),
                EventAggregator(client),
                FileAggregator(client),
                TimeSeriesAggregator(client),
                SequenceAggregator(client),
            ]
        }
        return self.create_profile_table(client, sheet=hierarchy)

    def create_initial_table(self, client: ToolkitClient) -> dict[tuple[AssetIndex, str], PendingCellValue]:
        table: dict[tuple[AssetIndex, str], PendingCellValue] = {}
        for index, aggregator in self.aggregators.items():
            asset_index = AssetIndex(aggregator=index, data_set_external_id=None, source=None)
            table[(asset_index, self.Columns.Resource)] = aggregator.display_name
            table[(asset_index, self.Columns.Count)] = WaitingAPICall
            table[(asset_index, self.Columns.DataSets)] = None
            table[(asset_index, self.Columns.DataSetCount)] = None
            table[(asset_index, self.Columns.Transformations)] = None
            table[(asset_index, self.Columns.RawTable)] = None
            table[(asset_index, self.Columns.RowCount)] = None
            table[(asset_index, self.Columns.ColumnCount)] = None
        return table

    def create_api_callable(self, row: AssetIndex, col: str, client: ToolkitClient) -> Callable:
        aggregator = self.aggregators[row.aggregator]
        if col == self.Columns.Count:
            return partial(aggregator.count, hierarchy=self.hierarchy)
        elif col == self.Columns.DataSets:
            return partial(aggregator.used_data_sets, hierarchy=self.hierarchy)
        elif col == self.Columns.DataSetCount:
            if row.data_set_external_id is None:
                raise ValueError(f"DataSet external ID is required for {row!s} in column {col}.")
            return partial(aggregator.count, data_set_external_id=row.data_set_external_id, hierarchy=self.hierarchy)
        elif col == self.Columns.Transformations:
            if row.data_set_external_id is None:
                raise ValueError(f"DataSet external ID is required for {row!s} in column {col}.")
            return partial(aggregator.used_transformations, data_set_external_ids=[row.data_set_external_id])
        elif col == self.Columns.ColumnCount:
            if row.source is None:
                raise ValueError(f"Database and table name are required for {row!s} in column {col}.")
            source = row.source
            return partial(
                client.raw.profile,
                database=source.db_name,
                table=source.table_name,
                limit=self.profile_row_limit,
                timeout_seconds=self.profile_timeout_seconds,
            )
        elif col == self.Columns.RowCount:
            if row.source is None:
                raise ValueError(f"Database and table name are required for {row!s} in column {col}.")
            source = row.source
            return partial(
                raw_row_count,
                client=client,
                raw_table_id=source,
            )
        raise ValueError(f"Unexpected API Call for row {row} and column {col}.")

    def format_result(self, result: object, row: AssetIndex, col: str) -> CellValue:
        if col == self.Columns.RowCount:
            if isinstance(result, int):
                if result == MAX_ROW_ITERATION_RUN_QUERY:
                    return f"≥{result:,}"
                else:
                    return result
            elif isinstance(result, str):
                return result
            return None
        elif isinstance(result, int | float | bool | str):
            return result
        elif col == self.Columns.DataSets:
            return result[0] if isinstance(result, list) and result and isinstance(result[0], str) else None
        elif col == self.Columns.Transformations:
            if isinstance(result, list) and len(result) > 0 and isinstance(result[0], Transformation):
                return f"{result[0].name} ({result[0].external_id})"
            return None
        elif col == self.Columns.ColumnCount:
            if isinstance(result, RawProfileResults):
                return result.column_count
            return None
        raise ValueError(f"unexpected result type {type(result)} for row {row!s} and column {col}.")

    def update_table(
        self,
        current_table: dict[tuple[AssetIndex, str], PendingCellValue],
        result: object,
        selected_row: AssetIndex,
        selected_col: str,
    ) -> dict[tuple[AssetIndex, str], PendingCellValue]:
        handlers: Mapping[
            str,
            Callable[
                [dict[tuple[AssetIndex, str], PendingCellValue], object, AssetIndex],
                dict[tuple[AssetIndex, str], PendingCellValue],
            ],
        ] = {
            self.Columns.Count: self._update_count,
            self.Columns.DataSets: self._update_datasets,
            self.Columns.DataSetCount: self._update_dataset_count,
            self.Columns.Transformations: self._update_transformations,
            self.Columns.ColumnCount: self._update_column_count,
        }
        handler = handlers.get(selected_col)
        if handler:
            return handler(current_table, result, selected_row)
        return current_table

    def _update_count(
        self,
        current_table: dict[tuple[AssetIndex, str], PendingCellValue],
        result: object,
        selected_row: AssetIndex,
    ) -> dict[tuple[AssetIndex, str], PendingCellValue]:
        new_table: dict[tuple[AssetIndex, str], PendingCellValue] = {}
        for (row, col), value in current_table.items():
            if row == selected_row and col == self.Columns.DataSets:
                new_table[(row, col)] = WaitingAPICall
            else:
                new_table[(row, col)] = value
        return new_table

    def _update_datasets(
        self,
        current_table: dict[tuple[AssetIndex, str], PendingCellValue],
        result: object,
        selected_row: AssetIndex,
    ) -> dict[tuple[AssetIndex, str], PendingCellValue]:
        if not (isinstance(result, list) and len(result) > 0 and all(isinstance(item, str) for item in result)):
            return current_table
        new_table: dict[tuple[AssetIndex, str], PendingCellValue] = {}
        for (row, col), value in current_table.items():
            if row != selected_row:
                new_table[(row, col)] = value
                continue
            for data_set in result:
                new_index = AssetIndex(
                    aggregator=selected_row.aggregator, source=selected_row.source, data_set_external_id=data_set
                )
                if col == self.Columns.DataSetCount:
                    new_table[(new_index, col)] = WaitingAPICall
                elif col == self.Columns.DataSets:
                    new_table[(new_index, col)] = data_set
                else:
                    new_table[(new_index, col)] = value
        return new_table

    def _update_dataset_count(
        self,
        current_table: dict[tuple[AssetIndex, str], PendingCellValue],
        result: object,
        selected_row: AssetIndex,
    ) -> dict[tuple[AssetIndex, str], PendingCellValue]:
        if not isinstance(result, int):
            return current_table
        new_table: dict[tuple[AssetIndex, str], PendingCellValue] = {}
        for (row, col), value in current_table.items():
            if row != selected_row:
                new_table[(row, col)] = value
                continue
            if col == self.Columns.Transformations:
                new_table[(row, col)] = WaitingAPICall
            else:
                new_table[(row, col)] = value
        return new_table

    def _update_transformations(
        self,
        current_table: dict[tuple[AssetIndex, str], PendingCellValue],
        result: object,
        selected_row: AssetIndex,
    ) -> dict[tuple[AssetIndex, str], PendingCellValue]:
        if not (
            isinstance(result, list) and len(result) > 0 and all(isinstance(item, Transformation) for item in result)
        ):
            return current_table
        sources_by_transformation_id = {
            transformation.id: [
                s for s in get_transformation_sources(transformation.query or "") if isinstance(s, RawTable)
            ]
            for transformation in result
        }
        new_table: dict[tuple[AssetIndex, str], PendingCellValue] = {}
        for (row, col), value in current_table.items():
            if row != selected_row:
                new_table[(row, col)] = value
                continue
            for transformation in result:
                sources = sources_by_transformation_id[transformation.id]
                if not sources:
                    new_table[(row, col)] = value
                    continue
                for source in sources:
                    new_index = AssetIndex(
                        aggregator=selected_row.aggregator,
                        data_set_external_id=selected_row.data_set_external_id,
                        source=source,
                    )
                    if col == self.Columns.RawTable:
                        new_table[(new_index, col)] = str(source)
                    elif col == self.Columns.RowCount:
                        # This will be updated by the ColumnCount API call which equals the /profiler/raw endpoint.
                        # If the profiles is not complete, we set it to WaitingAPICall to get the
                        # row count from the transformation preview.
                        new_table[(new_index, col)] = None
                    elif col == self.Columns.ColumnCount:
                        new_table[(new_index, col)] = WaitingAPICall
                    elif col == self.Columns.Transformations:
                        new_table[(new_index, col)] = f"{transformation.name} ({transformation.external_id})"
                    else:
                        new_table[(new_index, col)] = value
        return new_table

    def _update_column_count(
        self,
        current_table: dict[tuple[AssetIndex, str], PendingCellValue],
        result: object,
        selected_row: AssetIndex,
    ) -> dict[tuple[AssetIndex, str], PendingCellValue]:
        if not isinstance(result, RawProfileResults):
            return current_table
        new_table: dict[tuple[AssetIndex, str], PendingCellValue] = {}
        for (row, col), value in current_table.items():
            if row != selected_row:
                new_table[(row, col)] = value
                continue
            is_complete = result.is_complete and result.row_count < self.profile_row_limit
            if col == self.Columns.RowCount:
                # If the profile is complete, we can use the row count directly.
                # If not we set it to WaitingAPICall to get the row count from the transformation preview.
                new_table[(row, col)] = result.row_count if is_complete else WaitingAPICall
            elif col == self.Columns.ColumnCount:
                new_table[(row, col)] = result.column_count if is_complete else f"≥{result.column_count:,}"
            else:
                new_table[(row, col)] = value
        return new_table


class ProfileAssetCentricCommand(ProfileCommand[str]):
    def __init__(
        self,
        output_spreadsheet: Path | None = None,
        print_warning: bool = True,
        skip_tracking: bool = False,
        silent: bool = False,
    ) -> None:
        super().__init__(output_spreadsheet, print_warning, skip_tracking, silent)
        self.hierarchy: str | None = None
        self.table_title = "Asset Centric Profile"
        self.aggregators: dict[str, AssetCentricAggregator] = {}

    class Columns:
        Resource = "Resource"
        Count = "Count"
        MetadataKeyCount = "Metadata Key Count"
        LabelCount = "Label Count"
        Transformation = "Transformations"

    def asset_centric(
        self, client: ToolkitClient, hierarchy: str | None = None, select_all: bool = False, verbose: bool = False
    ) -> list[dict[str, CellValue]]:
        if hierarchy is None and not select_all:
            self.hierarchy = AssetInteractiveSelect(client, "profile").select_hierarchy(allow_empty=True)
            self._ask_store_file()
        else:
            self.hierarchy = hierarchy
        if self.hierarchy is not None:
            self.table_title = f"Asset Centric Profile: {self.hierarchy}"
        self.aggregators.update(
            {
                agg.display_name: agg
                for agg in [
                    AssetAggregator(client),
                    EventAggregator(client),
                    FileAggregator(client),
                    TimeSeriesAggregator(client),
                    SequenceAggregator(client),
                ]
            }
        )
        if self.hierarchy is None:
            # Relationship and Labels does not belong to a specific hierarchy
            self.aggregators.update(
                {
                    agg.display_name: agg
                    for agg in [
                        RelationshipAggregator(client),
                        LabelCountAggregator(client),
                    ]
                }
            )
        result = self.create_profile_table(client, sheet=self.table_title)
        if self.output_spreadsheet:
            for aggregator in self.aggregators.values():
                if isinstance(aggregator, AssetAggregator):
                    used_metadata_keys = aggregator.used_metadata_keys(hierarchy=self.hierarchy)
                    self._write_to_spreadsheet(
                        [{"Metadata Key": key, "Count": count} for key, count in used_metadata_keys],
                        ["Metadata Key", "Count"],
                        self.output_spreadsheet,
                        sheet=aggregator.display_name,
                    )
        return result

    def create_initial_table(self, client: ToolkitClient) -> dict[tuple[str, str], PendingCellValue]:
        table: dict[tuple[str, str], str | int | float | bool | None | WaitingAPICallClass] = {}
        for index, aggregator in self.aggregators.items():
            table[(index, self.Columns.Resource)] = aggregator.display_name
            table[(index, self.Columns.Count)] = WaitingAPICall
            # Metadata Key count is only valid if we aggregate for all resources or assets.
            # Events/Files/TimeSeries/Sequences do not have a rootId to filter on.
            if isinstance(aggregator, MetadataAggregator) and (
                isinstance(aggregator, AssetAggregator) or self.hierarchy is None
            ):
                table[(index, self.Columns.MetadataKeyCount)] = WaitingAPICall
            else:
                table[(index, self.Columns.MetadataKeyCount)] = None
            if isinstance(aggregator, LabelAggregator) and (
                isinstance(aggregator, AssetAggregator) or self.hierarchy is None
            ):
                table[(index, self.Columns.LabelCount)] = WaitingAPICall
            else:
                table[(index, self.Columns.LabelCount)] = None
            table[(index, self.Columns.Transformation)] = WaitingAPICall if self.hierarchy is None else None
        return table

    def create_api_callable(self, row: str, col: str, client: ToolkitClient) -> Callable:
        aggregator = self.aggregators[row]
        if col == self.Columns.Count:
            return partial(aggregator.count, hierarchy=self.hierarchy)
        elif col == self.Columns.MetadataKeyCount and isinstance(aggregator, MetadataAggregator):
            return aggregator.metadata_key_count
        elif col == self.Columns.LabelCount and isinstance(aggregator, LabelAggregator):
            return aggregator.label_count
        elif col == self.Columns.Transformation:
            return aggregator.transformation_count
        raise ValueError(f"Unknown column: {col} for row: {row}")


class ProfileTransformationCommand(ProfileCommand[str]):
    def __init__(
        self,
        output_spreadsheet: Path | None = None,
        print_warning: bool = True,
        skip_tracking: bool = False,
        silent: bool = False,
    ) -> None:
        super().__init__(output_spreadsheet, print_warning, skip_tracking, silent)
        self.table_title = "Transformation Profile"
        self.destination_type: AssetCentricDestinationType | None = None

    class Columns:
        Transformation = "Transformation"
        Source = "Sources"
        DestinationColumns = "Destination Columns"
        Destination = "Destination"
        ConflictMode = "Conflict Mode"
        IsPaused = "Is Paused"

    def transformation(
        self, client: ToolkitClient, destination_type: str | None = None, verbose: bool = False
    ) -> list[dict[str, CellValue]]:
        self.destination_type = AssetCentricDestinationSelect.get(destination_type)
        self.table_title = f"Transformation Profile destination: {self.destination_type}"
        return self.create_profile_table(client, sheet=self.destination_type)

    def create_initial_table(self, client: ToolkitClient) -> dict[tuple[str, str], PendingCellValue]:
        if self.destination_type is None:
            raise ToolkitValueError("Destination type must be set before calling create_initial_table.")
        iterable: Iterable[Transformation] = client.transformations.list(
            destination_type=self.destination_type, limit=-1
        )
        if self.destination_type == "assets":
            iterable = itertools.chain(iterable, client.transformations(destination_type="asset_hierarchy", limit=-1))
        table: dict[tuple[str, str], PendingCellValue] = {}
        for transformation in iterable:
            sources: list[SQLTable] = []
            destination_columns: list[str] = []
            if transformation.query:
                parser = SQLParser(transformation.query, operation="Profile transformations")
                sources = parser.sources
                destination_columns = parser.destination_columns
            index = str(transformation.id)
            table[(index, self.Columns.Transformation)] = transformation.name or transformation.external_id or "Unknown"
            table[(index, self.Columns.Source)] = ", ".join(map(str, sources))
            table[(index, self.Columns.DestinationColumns)] = (
                ", ".join(destination_columns) or None if destination_columns else None
            )
            table[(index, self.Columns.Destination)] = (
                transformation.destination.type or "Unknown" if transformation.destination else "Unknown"
            )
            table[(index, self.Columns.ConflictMode)] = transformation.conflict_mode or "Unknown"
            table[(index, self.Columns.IsPaused)] = (
                str(transformation.schedule.is_paused) if transformation.schedule else "No schedule"
            )
        return table

    def create_api_callable(self, row: str, col: str, client: ToolkitClient) -> Callable:
        raise NotImplementedError(f"{type(self).__name__} does not support API calls for {col} in row {row}.")


@dataclass(frozen=True)
class RawProfileIndex:
    raw_table: RawTable
    transformation_id: int | None = None


class ProfileRawCommand(ProfileCommand[RawProfileIndex]):
    class Columns:
        RAW = "Raw"
        Rows = "Rows"
        Columns = "Columns"
        Transformation = "Transformation"
        Destination = "Destination"
        ConflictMode = "ConflictMode"

    max_profile_raw_count = 10_000  # The number of rows to profile to get the number of columns.
    # The actual limit is 256 MB of data.
    # Ref https://github.com/cognitedata/profiler-api/blob/main/src/main/scala/com/cognite/raw_profiler/Profile.scala#L37
    profile_timeout_seconds = 60 * 4  # Timeout for the profiling operation in seconds,
    # This is the same ase the run/query maximum timeout.

    is_dynamic_table = True

    def __init__(
        self,
        output_spreadsheet: Path | None = None,
        print_warning: bool = True,
        skip_tracking: bool = False,
        silent: bool = False,
    ) -> None:
        super().__init__(output_spreadsheet, print_warning, skip_tracking, silent)
        self.table_title = "RAW Profile"
        self.destination_type: AssetCentricDestinationType | None = None
        self.client: ToolkitClient | None = None

    def raw(
        self,
        client: ToolkitClient,
        destination_type: str | None = None,
        verbose: bool = False,
    ) -> list[dict[str, CellValue]]:
        self.destination_type = AssetCentricDestinationSelect.get(destination_type)
        self.table_title = f"RAW Profile destination: {self.destination_type}"
        return self.create_profile_table(client, sheet=self.destination_type)

    def create_initial_table(self, client: ToolkitClient) -> dict[tuple[RawProfileIndex, str], PendingCellValue]:
        if self.destination_type is None:
            raise ToolkitValueError("Destination type must be set before calling create_initial_table.")
        table: dict[tuple[RawProfileIndex, str], PendingCellValue] = {}
        existing_tables = self._get_existing_tables(client)
        transformations_by_raw_table = self._get_transformations_by_raw_table(client, self.destination_type)
        for raw_id, transformations in transformations_by_raw_table.items():
            is_missing = raw_id not in existing_tables
            missing_str = " (missing)" if is_missing else ""
            for transformation in transformations:
                index = RawProfileIndex(raw_table=raw_id, transformation_id=transformation.id)
                table[(index, self.Columns.RAW)] = f"{raw_id!s}{missing_str}"
                if is_missing:
                    table[(index, self.Columns.Rows)] = "N/A"
                    table[(index, self.Columns.Columns)] = "N/A"
                else:
                    # First, we request the API to get the column count. This uses the /profiler/raw endpoint,
                    # and if that is a complete profile, it will update the row count as well. If it is not complete,
                    # the raw count will be obtained from the transformation/preview endpoint by setting the
                    # table[(index, self.Columns.Rows)] to WaitingAPICall.
                    table[(index, self.Columns.Rows)] = None
                    table[(index, self.Columns.Columns)] = WaitingAPICall
                table[(index, self.Columns.Transformation)] = f"{transformation.name} ({transformation.external_id})"
                table[(index, self.Columns.Destination)] = (
                    transformation.destination.type if transformation.destination else "Unknown"
                )
                table[(index, self.Columns.ConflictMode)] = transformation.conflict_mode
        return table

    def create_api_callable(self, row: RawProfileIndex, col: str, client: ToolkitClient) -> Callable:
        if col == self.Columns.Columns:
            return partial(
                client.raw.profile,
                database=row.raw_table.db_name,
                table=row.raw_table.table_name,
                limit=self.max_profile_raw_count,
                timeout_seconds=self.profile_timeout_seconds,
            )
        elif col == self.Columns.Rows:
            return partial(
                raw_row_count,
                client=client,
                raw_table_id=row.raw_table,
            )
        raise ValueError(f"There are no API calls for {row} in column {col}.")

    def format_result(self, result: object, row: RawProfileIndex, col: str) -> CellValue:
        if col == self.Columns.Rows:
            if isinstance(result, int):
                if result == MAX_ROW_ITERATION_RUN_QUERY:
                    return f"≥{result:,}"
                else:
                    return result
            elif isinstance(result, str):
                return result
            return None
        if isinstance(result, int | float | bool | str) or result is None:
            return result
        elif isinstance(result, RawProfileResults) and col == self.Columns.Columns:
            return result.column_count
        raise ValueError(f"Unknown result type: {type(result)} for {row!s} in column {col}.")

    def update_table(
        self,
        current_table: dict[tuple[RawProfileIndex, str], PendingCellValue],
        result: object,
        selected_row: RawProfileIndex,
        selected_col: str,
    ) -> dict[tuple[RawProfileIndex, str], PendingCellValue]:
        if not isinstance(result, RawProfileResults) or selected_col != self.Columns.Columns:
            return current_table
        is_complete = result.is_complete and result.row_count < self.max_profile_raw_count
        new_table: dict[tuple[RawProfileIndex, str], PendingCellValue] = {}
        for (row, col), value in current_table.items():
            if row == selected_row and col == self.Columns.Rows:
                new_table[(row, col)] = result.row_count if is_complete else WaitingAPICall
            elif row == selected_row and col == self.Columns.Columns:
                new_table[(row, col)] = result.column_count if is_complete else f"≥{result.column_count:,}"
            else:
                new_table[(row, col)] = value
        return new_table

    @classmethod
    def _get_existing_tables(cls, client: ToolkitClient) -> set[RawTable]:
        existing_tables: set[RawTable] = set()
        databases = client.raw.databases.list(limit=-1)
        for database in databases:
            if database.name is None:
                continue
            tables = client.raw.tables.list(db_name=database.name, limit=-1)
            for table in tables:
                if table.name is None:
                    continue
                existing_tables.add(RawTable(db_name=database.name, table_name=table.name))
        return existing_tables

    @classmethod
    def _get_transformations_by_raw_table(
        cls, client: ToolkitClient, destination_type: str
    ) -> dict[RawTable, list[Transformation]]:
        transformations = client.transformations.list(destination_type=destination_type)
        if destination_type == "assets":
            transformations.extend(client.transformations.list(destination_type="asset_hierarchy"))
        transformations_by_raw_table: dict[RawTable, list[Transformation]] = defaultdict(list)
        for transformation in transformations:
            if transformation.query is None:
                # No query means no source table.
                continue
            sources = get_transformation_sources(transformation.query)
            for source in sources:
                if isinstance(source, RawTable):
                    transformations_by_raw_table[source].append(transformation)
        return transformations_by_raw_table
