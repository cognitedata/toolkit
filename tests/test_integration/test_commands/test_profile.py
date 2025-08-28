from pathlib import Path

import pytest
from cognite.client.data_classes import Asset, Transformation
from openpyxl import load_workbook

from cognite_toolkit._cdf_tk.client import ToolkitClient
from cognite_toolkit._cdf_tk.commands import (
    ProfileAssetCentricCommand,
    ProfileAssetCommand,
    ProfileRawCommand,
    ProfileTransformationCommand,
)
from tests.test_integration.constants import (
    ASSET_COUNT,
    ASSET_DATASET,
    ASSET_TABLE,
    EVENT_COUNT,
    EVENT_DATASET,
    EVENT_TABLE,
    FILE_COUNT,
    FILE_DATASET,
    FILE_TABLE,
    SEQUENCE_COUNT,
    SEQUENCE_DATASET,
    SEQUENCE_TABLE,
    TIMESERIES_COUNT,
    TIMESERIES_DATASET,
    TIMESERIES_TABLE,
)


class TestProfileAssetCommand:
    @pytest.mark.usefixtures("disable_throttler")
    def test_profile_asset_hierarchy(
        self,
        toolkit_client: ToolkitClient,
        aggregator_root_asset: Asset,
        aggregator_raw_db: str,
        aggregator_assets: Transformation,
        aggregator_events: Transformation,
        aggregator_files: Transformation,
        aggregator_time_series: Transformation,
        aggregator_sequences: Transformation,
    ) -> None:
        profile_row_limit = TIMESERIES_COUNT - 1  # Force fallaback to using the preview endpoint for row count
        results = ProfileAssetCommand().assets(
            toolkit_client, aggregator_root_asset.external_id, profile_row_limit=profile_row_limit
        )
        for row in results:
            for cell in row.values():
                if "CogniteAPIError" in str(cell):
                    pytest.skip("Skipping test due to intermediate CogniteAPI error.")

        columns = ProfileAssetCommand.Columns
        assert (
            results
            == [
                {
                    columns.Resource: resource,
                    columns.Count: count,
                    columns.DataSets: dataset,
                    columns.DataSetCount: count,
                    columns.Transformations: f"{transformation.name} ({transformation.external_id})",
                    columns.RawTable: f"{aggregator_raw_db}.{raw_table}",
                    columns.RowCount: row_count,
                    columns.ColumnCount: column_count,
                }
                for resource, count, dataset, transformation, raw_table, row_count, column_count in [
                    (
                        "Assets",
                        ASSET_COUNT,
                        ASSET_DATASET,
                        aggregator_assets,
                        ASSET_TABLE,
                        ASSET_COUNT - 1,
                        4,
                    ),  # -1 root asset is not in the table
                    (
                        "Events",
                        EVENT_COUNT,
                        EVENT_DATASET,
                        aggregator_events,
                        EVENT_TABLE,
                        EVENT_COUNT,
                        5,
                    ),
                    (
                        "Files",
                        FILE_COUNT,
                        FILE_DATASET,
                        aggregator_files,
                        FILE_TABLE,
                        FILE_COUNT,
                        4,
                    ),
                    (
                        "TimeSeries",
                        TIMESERIES_COUNT,
                        TIMESERIES_DATASET,
                        aggregator_time_series,
                        TIMESERIES_TABLE,
                        TIMESERIES_COUNT,
                        "≥5",  # Force fallback to using the preview endpoint for row count, then we get "≥5" as the column count
                    ),
                    (
                        "Sequences",
                        SEQUENCE_COUNT,
                        SEQUENCE_DATASET,
                        aggregator_sequences,
                        SEQUENCE_TABLE,
                        SEQUENCE_COUNT,
                        4,
                    ),
                ]
            ]
        )


class TestProfileAssetCentric:
    def test_profile_assent_centric(self, toolkit_client: ToolkitClient, monkeypatch, tmp_path: Path) -> None:
        output_spreadsheet = tmp_path / "asset_centric_profile.xlsx"
        cmd = ProfileAssetCentricCommand(output_spreadsheet)
        results = cmd.asset_centric(toolkit_client, select_all=True, verbose=False)

        assert len(results) == 7
        assert {item["Resource"] for item in results} == {
            "Assets",
            "Events",
            "Files",
            "TimeSeries",
            "Sequences",
            "Relationships",
            "Labels",
        }
        total_count = sum(item["Count"] for item in results)
        assert total_count > 0
        total_metadata_count = 0
        for item in results:
            metadata_count = item.get(ProfileAssetCentricCommand.Columns.MetadataKeyCount)
            if not metadata_count:
                continue
            total_metadata_count += metadata_count
        assert total_metadata_count > 0

        workbook = load_workbook(output_spreadsheet)
        assert cmd.table_title in workbook.sheetnames
        sheet = workbook[cmd.table_title]
        assert sheet.max_row == len(results) + 1  # +1 for header row

    def test_profile_asset_centric_hierarchy(
        self, toolkit_client: ToolkitClient, aggregator_root_asset: Asset, tmp_path: Path
    ) -> None:
        output_spreadsheet = tmp_path / "asset_centric_hierarchy_profile.xlsx"
        cmd = ProfileAssetCentricCommand(silent=True, output_spreadsheet=output_spreadsheet)
        results = cmd.asset_centric(toolkit_client, hierarchy=aggregator_root_asset.external_id, verbose=False)
        actual = [{"Resource": row["Resource"], "Count": row["Count"]} for row in results]
        expected = [
            {"Resource": "Assets", "Count": ASSET_COUNT},
            {"Resource": "Events", "Count": EVENT_COUNT},
            {"Resource": "Files", "Count": FILE_COUNT},
            {"Resource": "TimeSeries", "Count": TIMESERIES_COUNT},
            {"Resource": "Sequences", "Count": SEQUENCE_COUNT},
        ]
        assert actual == expected, f"Expected {expected}, but got {actual}"

        workbook = load_workbook(output_spreadsheet)
        assert "Assets" in workbook.sheetnames
        sheet = workbook["Assets"]
        assert [col.value for (col, *_) in sheet.columns] == ["Metadata Key", "Count"]


class TestProfileTransformationCommand:
    def test_profile_transformation(
        self, toolkit_client: ToolkitClient, aggregator_assets: Transformation, aggregator_raw_db: str
    ) -> None:
        results = ProfileTransformationCommand().transformation(toolkit_client, "assets")
        columns = ProfileTransformationCommand.Columns
        search_rows = [row for row in results if row[columns.Transformation] == aggregator_assets.name]
        assert len(search_rows) == 1, "Expected exactly one row for the transformation"
        actual_row = search_rows[0]
        assert actual_row == {
            columns.Transformation: aggregator_assets.name,
            columns.Source: f"{aggregator_raw_db}.{ASSET_TABLE}",
            columns.DestinationColumns: "name, externalId, dataSetId, parentExternalId",
            columns.Destination: "assets",
            columns.ConflictMode: aggregator_assets.conflict_mode,
            columns.IsPaused: "No schedule",
        }


class TestProfileRawCommand:
    def test_profile_raw(
        self, toolkit_client: ToolkitClient, aggregator_raw_db: str, aggregator_assets: Transformation
    ) -> None:
        transformation = aggregator_assets
        results = ProfileRawCommand().raw(toolkit_client, destination_type="assets", verbose=False)
        columns = ProfileRawCommand.Columns
        search_rows = [row for row in results if row[columns.RAW] == f"{aggregator_raw_db}.{ASSET_TABLE}"]
        assert len(search_rows) == 1, f"Expected exactly one row for the raw table {ASSET_TABLE}."
        actual_row = search_rows[0]
        assert actual_row == {
            columns.RAW: f"{aggregator_raw_db}.{ASSET_TABLE}",
            columns.Rows: ASSET_COUNT - 1,  # -1 root asset is not in the RAW table.
            columns.Columns: 4,
            columns.Transformation: f"{transformation.name} ({transformation.external_id})",
            columns.Destination: "assets",
            columns.ConflictMode: transformation.conflict_mode,
        }
